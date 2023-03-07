import openpyxl as xl
from openpyxl.chart import BarChart, Reference


# Define a function to use DRY rules
def process_workbook(filename):
    wb = xl.load_workbook(filename)

    # Get the sheet using sheet name
    sheet = wb['Sheet1']

    # ways to get the cells'
    # cell = sheet['a1']
    # cell = sheet.cell(row, column)

    # get the max number of rows (count)
    max_rows = sheet.max_row

    # Added the column name
    sheet.cell(1, 4).value = "updated_price"

    for row in range(2, max_rows + 1):
        cell = sheet.cell(row, 3)  # Get the cell number as passing the current row and 3rd column
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    # setup values to passing it in the chart
    values = Reference(sheet,
                       min_row=2,
                       max_row=max_rows,
                       min_col=4,
                       max_col=4)
    chart = BarChart()
    chart.add_data(values)  # adding the data in the chart
    sheet.add_chart(chart, 'e2')  # Adding the chart in the sheet at e2

    wb.save(file_name)


# Calling the above function using the below file
file_name = 'transactions_new.xlsx'
process_workbook(file_name)
